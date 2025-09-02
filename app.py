import streamlit as st
import tempfile
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

st.set_page_config(page_title="Unir Excels", page_icon="📊", layout="centered")

st.title("📊 Unir Excels en Plantilla (.xlsm)")

# Subida de archivos
archivos = st.file_uploader("📂 Sube tus archivos Excel", type=["xlsx", "xlsm"], accept_multiple_files=True)

if archivos:
    plantilla = st.file_uploader("📂 Sube tu plantilla base", type=["xlsx", "xlsm"])

    if plantilla:
        if st.button("🚀 Procesar"):
            progress = st.progress(0)

            # Crear carpeta temporal
            temp_dir = tempfile.mkdtemp()

            # Guardar plantilla con su extensión original
            plantilla_path = os.path.join(temp_dir, plantilla.name)
            with open(plantilla_path, "wb") as f:
                f.write(plantilla.read())

            # Abrir plantilla
            wb_destino = load_workbook(plantilla_path, keep_vba=True)
            hoja_destino = wb_destino.active

            fila_destino = 2  # empezar en la fila 2
            total = len(archivos)

            # Definir estilo de fuente
            estilo_fuente = Font(name="Calibri", size=10)

            for i, archivo in enumerate(archivos):
                # Guardar archivo origen
                file_path = os.path.join(temp_dir, archivo.name)
                with open(file_path, "wb") as f:
                    f.write(archivo.read())

                wb_origen = load_workbook(file_path, data_only=True)
                hoja_origen = wb_origen.active

                # Iterar filas (ignorando cabecera)
                for fila in hoja_origen.iter_rows(min_row=2, values_only=True):
                    if fila is None:
                        continue

                    valores = [celda for celda in fila if celda not in (None, "")]
                    if not valores:
                        continue

                    if len(valores) > 10:
                        for col, valor in enumerate(fila, start=1):
                            celda_destino = hoja_destino.cell(row=fila_destino, column=col, value=valor)
                            # Aplicar estilo de fuente
                            celda_destino.font = estilo_fuente

                            # Si es fecha → aplicar formato fecha corta
                            if isinstance(valor, datetime):
                                celda_destino.number_format = "dd/mm/yyyy"

                        fila_destino += 1

                # Escribir referencia del archivo
                hoja_destino.cell(row=fila_destino, column=1, value=f"📌 Datos provenientes de: {archivo.name}")
                hoja_destino.cell(row=fila_destino, column=1).font = estilo_fuente
                fila_destino += 2

                wb_origen.close()

                # Actualizar barra progreso
                progress.progress((i + 1) / total)

            # Guardar resultado
            if plantilla.name.endswith(".xlsm"):
                resultado_path = os.path.join(temp_dir, "resultado.xlsm")
            else:
                resultado_path = os.path.join(temp_dir, "resultado.xlsx")

            wb_destino.save(resultado_path)
            wb_destino.close()

            # Botón descargar
            with open(resultado_path, "rb") as f:
                st.download_button(
                    label="📥 Descargar Resultado",
                    data=f,
                    file_name=os.path.basename(resultado_path),
                    mime="application/vnd.ms-excel.sheet.macroEnabled.12" if resultado_path.endswith(".xlsm")
                         else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
